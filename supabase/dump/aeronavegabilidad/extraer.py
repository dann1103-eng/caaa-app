#!/usr/bin/env python3
"""
Lee la documentacion de aeronavegabilidad y la normaliza a JSON.

Paso 1 de 2 (el 2 es cargar.js). Va en Python porque NO hay modulo de Excel para
Node en este proyecto — misma razon que la carga del inventario OMA (CLAUDE.md §16).

Fuentes:
  - ADs de 4 aviones: Excel de 3 hojas (AVION / MOTOR / HELICE).
  - ADs del YS-127-P: _transcripcion_OCR.json (sus PDF venian escaneados).
  - Vida limite del YS-334-PE: .docx con tablas reales.
  - Vida limite del YS-127-P: el mismo JSON del OCR.

NO se leen las listas de vida limite del YS-270-P ni del YS-333-PE, que solo
existen como PDF: la extraccion de texto corre la columna TIME un renglon hacia
arriba (el item 13 "VACUM SYSTEM FILTER" sale con "12 Years", que es el
intervalo del 12). El intervalo es lo que dispara el vencimiento, asi que
importarlos asi meteria numeros equivocados en un registro que la AAC audita.
Se reportan como pendientes de dictado y punto.

Uso:  python extraer.py
Sale: aeronavegabilidad.json
"""
import json, re, sys, unicodedata
from pathlib import Path
from collections import defaultdict

RAIZ = Path(__file__).resolve().parents[3]
DOCS = RAIZ / "docs" / "formatos-aac" / "aeronavegabilidad"
SALIDA = Path(__file__).resolve().parent / "aeronavegabilidad.json"

# La carpeta se llama "Docs. YS-270-P" pero la matricula real es YS-270-PE
# (error historico ya corregido en la BD, CLAUDE.md §7).
MATRICULA = {
    "YS-127-P": "YS-127-P", "YS-270-P": "YS-270-PE", "YS-333-PE": "YS-333-PE",
    "YS-334-PE": "YS-334-PE", "YS-361-PE": "YS-361-PE",
}
LIBRO = {"AVION": "CELULA", "AERONAVE": "CELULA", "MOTOR": "MOTOR", "HELICE": "HELICE"}

problemas = []
def aviso(avion, libro, msg):
    problemas.append({"avion": avion, "libro": libro, "detalle": msg})


def sinacentos(s):
    return "".join(c for c in unicodedata.normalize("NFD", str(s or "")) if unicodedata.category(c) != "Mn")


def txt(v):
    return re.sub(r"\s+", " ", sinacentos(v).replace("\n", " ")).strip()


def es_na(v):
    return txt(v).upper() in ("", "N/A", "NA", "N.A", "-", "—")


def a_fecha(v):
    """Excel devuelve datetime; el OCR devuelve 'DD/MM/YYYY' o 'D/M/YYYY'."""
    if v is None or es_na(v):
        return None
    if hasattr(v, "year"):
        return f"{v.year:04d}-{v.month:02d}-{v.day:02d}"
    s = txt(v)
    m = re.match(r"^(\d{1,2})[/-](\d{1,2})[/-](\d{4})$", s)
    if m:
        d, mes, a = int(m.group(1)), int(m.group(2)), int(m.group(3))
        if 1 <= mes <= 12 and 1 <= d <= 31:
            return f"{a:04d}-{mes:02d}-{d:02d}"
    m = re.match(r"^(\d{4})-(\d{2})-(\d{2})", s)
    return m.group(0) if m else None


def a_horas(v, avion=None, libro=None, ref=None):
    """
    Horas en ESCALA DE LIBRO, tal como estan en el papel. cargar.js les resta el
    tac_offset. Devuelve None si no se puede leer con confianza.
    """
    if v is None or es_na(v):
        return None
    if hasattr(v, "year"):
        # Una FECHA en la columna de horas: el AD 2018-07-03 del 127 tiene
        # "31/5/1920" ahi, autoformato de Excel. El TAC real se perdio.
        aviso(avion, libro, f"{ref}: la columna de TAC trae una fecha ({v}); el TAC se perdio")
        return None
    s = txt(v).upper().replace("TAC", "").replace("HRS", "").replace("HR", "").strip()
    # Notacion h:mm (9,844:42 = 9844.70 hrs, NO 9844.42). Vive en la vida limite
    # del 127 mezclada con decimales en el mismo documento.
    m = re.match(r"^([\d,]+):(\d{1,2})$", s)
    if m:
        return round(float(m.group(1).replace(",", "")) + int(m.group(2)) / 60.0, 2)
    s = s.replace(",", "")
    # Coma decimal suelta ya normalizada arriba; queda el numero.
    m = re.match(r"^-?\d+(\.\d+)?$", s)
    if m:
        return round(float(s), 2)
    if s:
        aviso(avion, libro, f"{ref}: no pude leer '{txt(v)}' como horas")
    return None


def intervalo_de(texto):
    """
    La columna RECURRENTE / TIME. Devuelve (horas, dias, recurrente, nota).

    Formas reales encontradas en los 5 aviones:
      "SI" (28 veces, sin decir cada cuanto)  ·  "TAC. 100"  ·  "Cada 100 HRS"
      "TAC 5,000.0"  ·  "O/C" (on condition)  ·  "2,400 Hrs"  ·  "12Yrs" / "7 YRS"
    """
    s = txt(texto).upper()
    if s in ("", "N/A", "NA", "NO"):
        return None, None, False, None
    if s in ("O/C", "OC", "ON CONDITION"):
        return None, None, True, "On condition (O/C): se revisa por estado, no por intervalo"
    horas = dias = None
    m = re.search(r"([\d,]+(?:\.\d+)?)\s*(?:HRS?|HORAS?)\b", s) or re.search(r"TAC\.?\s*([\d,]+(?:\.\d+)?)", s)
    if m:
        horas = round(float(m.group(1).replace(",", "")), 2)
    m = re.search(r"([\d,]+(?:\.\d+)?)\s*(?:YRS?|YEARS?|ANOS?)\b", s)
    if m:
        dias = int(round(float(m.group(1).replace(",", "")) * 365))
    if horas is None and dias is None and re.fullmatch(r"[\d,]+(\.\d+)?", s):
        horas = round(float(s.replace(",", "")), 2)
    return horas, dias, True, None


def aplica_de(observaciones):
    """N/A por serie, por modelo, por P/N, o no instalado ⇒ no aplica al avion."""
    s = txt(observaciones).upper()
    return not (s.startswith("N/A") or s.startswith("NO APLICA") or "NO INSTALADO" in s)


# ── ADs desde los Excel ───────────────────────────────────────────────────
def leer_ads_excel(ruta, avion):
    import openpyxl
    filas = []
    wb = openpyxl.load_workbook(ruta, data_only=True)
    for ws in wb.worksheets:
        hoja = txt(ws.title).upper()
        libro = next((v for k, v in LIBRO.items() if k in hoja), None)
        if not libro:
            aviso(avion, hoja, "hoja sin libro reconocible, se omite")
            continue

        # El encabezado no esta siempre en la misma fila y las columnas cambian
        # de avion en avion. Se busca la fila que tiene "AD" en la primera celda.
        hdr = None
        for fila in ws.iter_rows(min_row=1, max_row=10):
            if re.search(r"\bAD", txt(fila[0].value).upper()):
                hdr = fila[0].row
                break
        if hdr is None:
            aviso(avion, libro, "no encontre la fila de encabezado, hoja omitida")
            continue

        cols = [txt(c.value).upper() for c in ws[hdr]]
        def col(*pats):
            for i, c in enumerate(cols):
                if any(p in c for p in pats):
                    return i
            return None
        # "FECHA Y/O TACOMETRO" del 334 es una sola columna: cae en fecha y el
        # TAC queda None, que es la verdad — no hay dos datos separados ahi.
        i_sb, i_desc = col("S.B", "SB"), col("DESCRIPCION")
        i_obs, i_fec = col("OBSERV"), col("FECHA")
        i_tac, i_una = col("TAC", "TACOMETRO"), col("UNA VEZ")
        i_rec, i_pro = col("RECURRENT"), col("PROXIMA")
        if i_tac is not None and i_tac == i_fec:
            i_tac = None

        for fila in ws.iter_rows(min_row=hdr + 1, max_row=ws.max_row):
            v = [c.value for c in fila]
            ref = txt(v[0]) if v else ""
            if not ref or not re.match(r"^\d{2,4}-\d{2}-\d{2}", ref):
                continue  # firmas, pies de pagina, filas vacias
            def g(i):
                return v[i] if (i is not None and i < len(v)) else None
            obs = txt(g(i_obs))
            h, d, rec, nota = intervalo_de(g(i_rec))
            if txt(g(i_una)).upper() == "SI" and rec:
                nota = ((nota + " · ") if nota else "") + "el papel marca UNA VEZ y RECURRENTE a la vez"
            filas.append({
                "libro": libro, "referencia": ref, "sb": txt(g(i_sb)) or None,
                "descripcion": txt(g(i_desc)) or ref, "observaciones": obs or None,
                "aplica": aplica_de(obs),
                "ultima_fecha": a_fecha(g(i_fec)),
                "ultima_horas_libro": a_horas(g(i_tac), avion, libro, ref),
                "recurrente": rec, "intervalo_horas": h, "intervalo_dias": d,
                "proxima_horas_libro": a_horas(g(i_pro), avion, libro, ref),
                "nota": nota, "origen": "EXCEL_2026",
            })
    return filas


# ── ADs y vida limite del 127, desde el JSON del OCR ──────────────────────
def leer_127(ruta):
    d = json.loads(ruta.read_text(encoding="utf-8"))
    filas = []
    for hoja, libro in (("AVION", "CELULA"), ("MOTOR", "MOTOR"), ("HELICE", "HELICE")):
        # ⚠️ Son OCHO columnas, no nueve: [ad, sb, descripcion, fecha, tac,
        # una_vez, recurrente, proxima]. La columna OBSERVACIONES del papel no
        # esta en las filas porque dice lo mismo en las 66 ("SE INSPECCIONO
        # SEGUN AD") y se omitio para no repetirla. Leer f[3] como
        # observaciones corre todo un lugar y deja 34 ADs marcados como que no
        # aplican al avion.
        obs = d["ads"]["_observacion_comun_papel"]
        for f in d["ads"][hoja]["filas"]:
            ref, sb, desc, fec, tac, una, rec_txt = f[0], f[1], f[2], f[3], f[4], f[5], f[6]
            h, dd, rec, nota = intervalo_de(rec_txt)
            if txt(una).upper() == "SI" and rec:
                nota = ((nota + " · ") if nota else "") + "el papel marca UNA VEZ y RECURRENTE a la vez"
            filas.append({
                "libro": libro, "referencia": ref, "sb": sb or None, "descripcion": desc,
                "observaciones": obs or None, "aplica": aplica_de(obs),
                "ultima_fecha": a_fecha(fec),
                "ultima_horas_libro": a_horas(tac, "YS-127-P", libro, ref),
                "recurrente": rec, "intervalo_horas": h, "intervalo_dias": dd,
                "proxima_horas_libro": None, "nota": nota, "origen": "OCR_2026",
            })
    vl = [_vl_fila("YS-127-P", r[0], r[1], r[2], r[3], r[4], r[5], "OCR_2026")
          for r in d["vida_limite"]["filas"]]
    return filas, agrupar_doble_base(vl)


def _vl_fila(avion, n, item, time, fecha, hour, next_due, origen):
    h, dd, _, nota = intervalo_de(time)
    # Los ADs vienen en TRES hojas (avion / motor / helice) y por eso se agrupan
    # por libro. La vida limite es UNA sola lista para todo el avion: su
    # cabecera cubre las tres partes juntas. Partirla por libro seria inventar
    # una division que el papel no tiene, asi que va sin libro y en el orden en
    # que esta escrita (orden_papel).
    return {
        "libro": None,
        "referencia": None, "sb": None, "descripcion": txt(item),
        "observaciones": None, "aplica": True,
        "ultima_fecha": a_fecha(fecha),
        "ultima_horas_libro": a_horas(hour, avion, "VIDA_LIMITE", txt(item)[:30]),
        "recurrente": True, "intervalo_horas": h, "intervalo_dias": dd,
        "proxima_horas_libro": a_horas(next_due, avion, "VIDA_LIMITE", txt(item)[:30]),
        "nota": nota, "orden_papel": txt(n), "origen": origen,
    }


def agrupar_doble_base(filas):
    """
    Un item con dos vencimientos (2,000 Hrs Y 12 Yrs) viene en DOS renglones
    identicos salvo la columna TIME. Se guarda como UNA fila con los dos
    intervalos: asi se cumple una vez y alerta una vez. La pantalla lo vuelve a
    mostrar como dos renglones (decision de Daniel: fidelidad renglon por
    renglon en lo que se ve, un solo registro en lo que se guarda).
    """
    fusion = {}
    for f in filas:
        k = (f.get("orden_papel"), f["descripcion"].upper())
        if k in fusion:
            g = fusion[k]
            g["intervalo_horas"] = g["intervalo_horas"] or f["intervalo_horas"]
            g["intervalo_dias"] = g["intervalo_dias"] or f["intervalo_dias"]
            g["doble_base"] = True
        else:
            fusion[k] = dict(f, doble_base=False)
    return list(fusion.values())


def leer_vl_docx(ruta, avion):
    import docx
    d = docx.Document(ruta)
    filas = []
    tabla = max(d.tables, key=lambda t: len(t.rows))
    for r in tabla.rows[1:]:
        c = [txt(x.text) for x in r.cells]
        if len(c) < 7 or not c[1]:
            continue
        filas.append(_vl_fila(avion, c[0], c[1], c[2], c[3], c[4], c[5], "EXCEL_2026"))
    return agrupar_doble_base(filas)


# ── Main ──────────────────────────────────────────────────────────────────
def main():
    if not DOCS.is_dir():
        sys.exit(f"No encuentro {DOCS}")
    aviones = {}
    for carpeta in sorted(DOCS.glob("Docs. *")):
        clave = carpeta.name.replace("Docs. ", "").strip()
        avion = MATRICULA.get(clave, clave)
        ads, vida = [], []

        ocr = carpeta / "_transcripcion_OCR.json"
        if ocr.exists():
            ads, vida = leer_127(ocr)
        else:
            xl = [p for p in carpeta.glob("*.xlsx") if re.search(r"ads?\b", p.name, re.I)]
            if xl:
                ads = leer_ads_excel(xl[0], avion)
            else:
                aviso(avion, None, "sin lista de ADs legible")

            dx = list(carpeta.glob("*vida limite*.docx"))
            if dx:
                vida = leer_vl_docx(dx[0], avion)
            elif list(carpeta.glob("*vida limite*.pdf")) or list(carpeta.glob("*Vida Limite*.pdf")):
                aviso(avion, "VIDA_LIMITE",
                      "la lista solo existe como PDF y su extraccion corre la columna TIME "
                      "un renglon hacia arriba; NO se importa. Hay que dictarla del papel.")
            else:
                aviso(avion, "VIDA_LIMITE", "este avion no tiene lista de vida limite entregada")

        aviones[avion] = {"ads": ads, "vida_limite": vida}

    resumen = {}
    for av, d in aviones.items():
        rec = [f for f in d["ads"] if f["recurrente"]]
        resumen[av] = {
            "ads": len(d["ads"]),
            "ads_aplican": sum(1 for f in d["ads"] if f["aplica"]),
            "ads_recurrentes": len(rec),
            "ads_recurrentes_sin_intervalo": sum(
                1 for f in rec if f["intervalo_horas"] is None and f["intervalo_dias"] is None),
            "vida_limite": len(d["vida_limite"]),
            "vida_limite_doble_base": sum(1 for f in d["vida_limite"] if f.get("doble_base")),
        }

    SALIDA.write_text(json.dumps(
        {"aviones": aviones, "resumen": resumen, "problemas": problemas},
        ensure_ascii=False, indent=1), encoding="utf-8")

    print(f"{'AVION':<12}{'ADs':>6}{'aplican':>9}{'recurr':>8}{'sin int.':>10}{'vida lim.':>11}{'doble base':>12}")
    tot = defaultdict(int)
    for av, r in sorted(resumen.items()):
        print(f"{av:<12}{r['ads']:>6}{r['ads_aplican']:>9}{r['ads_recurrentes']:>8}"
              f"{r['ads_recurrentes_sin_intervalo']:>10}{r['vida_limite']:>11}{r['vida_limite_doble_base']:>12}")
        for k, v in r.items():
            tot[k] += v
    print("-" * 68)
    print(f"{'TOTAL':<12}{tot['ads']:>6}{tot['ads_aplican']:>9}{tot['ads_recurrentes']:>8}"
          f"{tot['ads_recurrentes_sin_intervalo']:>10}{tot['vida_limite']:>11}{tot['vida_limite_doble_base']:>12}")
    print(f"\n{len(problemas)} avisos -> {SALIDA.name} (clave \"problemas\")")


if __name__ == "__main__":
    main()
