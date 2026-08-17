# -*- coding: utf-8 -*-
"""
Paso 1 de la carga del inventario de la bodega OMA: Excel -> JSON normalizado.

    python extraer_excel.py "<ruta al .xlsx>" inventario_oma.json

Va en dos pasos (Python lee el Excel, Node carga a Supabase) porque en este
repo no hay módulo de Excel para Node — la lección de la carga del programa
semanal (CLAUDE.md §16).

Lo que arregla acá, con el detalle en la lista `problemas` del JSON:

  · La llave del movimiento pasa a ser el CÓDIGO del ítem. El Excel cruzaba por
    descripción + número de parte, y por eso 26 líneas de entrada y 9 de salida
    eran invisibles para el stock.
  · Correlativos rotos: FA-0000025-2026 (7 dígitos) es el mismo documento que
    FA-00025-2026, y FA-00019-2027 tiene el año mal tecleado. Se agrupan por
    (tipo, año de la FECHA, número), así los dos casos se fusionan solos.
  · Clasificación, unidad de medida y S/N embebido en la descripción.
  · La matrícula de la aeronave se extrae del comentario de la salida.

Spec: docs/superpowers/specs/2026-08-17-inventario-taller-design.md
"""
import json
import re
import sys
from collections import defaultdict

import openpyxl

# Matrículas tal como aparecen en los comentarios -> matrícula real de la BD.
# El Excel arrastra typos históricos (YS-270-P por YS-270-PE, ver CLAUDE.md §16).
ALIAS_MATRICULA = {
    "YS-334-P": "YS-334-PE",
    "YS-270-P": "YS-270-PE",
    "YS-127-PE": "YS-127-P",
    "YS-155-P": "YS-155-PE",
}
# Aviones de terceros: la OMA les da mantenimiento pero no son de la escuela.
EXTERNAS = {"YS-361-PE", "YS-243-P", "YS-22-C"}

UNIDADES = {
    "QT": "QT", "QTO": "QT", "QTS": "QT", "QUART": "QT",
    "FT": "FT", "PIE": "FT", "PIES": "FT", "FEET": "FT",
    "GAL": "GAL", "GALON": "GAL", "GALLON": "GAL",
    "KIT": "KIT", "JGO": "JGO", "JUEGO": "JGO", "SET": "JGO",
    "LB": "LB", "LBS": "LB", "LIBRA": "LB",
}
CLASIFICACION = {
    "ACEITES": "ACEITE",
    "INSTUMENTOS": "INSTRUMENTOS",
    "INSTRUMENTO": "INSTRUMENTOS",
    "ROTABLE": "ROTABLES",
    "TOOL": "HERRAMIENTA",
    "ELECTRONICO": "ELECTRICA",
}


def txt(v):
    return "" if v is None else str(v).strip()


def norm(v):
    return re.sub(r"\s+", " ", txt(v)).upper()


def unidad(v):
    return UNIDADES.get(norm(v), "UN")


def clasificacion(v):
    c = norm(v)
    if not c or c.isdigit():        # un código que se coló de clasificación
        return None
    return CLASIFICACION.get(c, c)


def extraer_serie(desc):
    """'VOR S/N 2253' -> ('VOR', '2253')."""
    d = txt(desc)
    m = re.match(r"^(.*?)\s*S/N\s*(\S+)\s*(.*)$", d, re.I)
    if not m:
        return d, None
    limpia = re.sub(r"\s+", " ", f"{m.group(1)} {m.group(3)}").strip()
    return (limpia or d), m.group(2)


def matricula(comentario):
    """Saca la matrícula del comentario libre de la salida."""
    c = norm(comentario)
    m = re.search(r"YS[\s-]*(\d{2,3})[\s-]*([A-Z]{1,2})?", c)
    if not m:
        return None
    base = f"YS-{m.group(1)}-{m.group(2) or ''}".rstrip("-")
    return ALIAS_MATRICULA.get(base, base)


def cantidad(v):
    """
    Cantidad tolerante. En el Excel hay celdas como '1FT' o '2´.4´´' (pies y
    pulgadas de manguera). Se toma el número de adelante para no perder el
    movimiento; el texto original queda en la nota del renglón.
    """
    if isinstance(v, (int, float)):
        return float(v), None
    s = txt(v)
    m = re.match(r"^\s*(\d+(?:[.,]\d+)?)", s)
    if not m:
        return None, None
    return float(m.group(1).replace(",", ".")), s


def filas(ws, col_desc, col_cant):
    """Filas con datos reales. El Excel tiene cientos de filas plantilla con '-'."""
    out = []
    for r in range(8, ws.max_row + 1):
        d = ws.cell(r, col_desc).value
        q = ws.cell(r, col_cant).value
        if txt(d) in ("", "-") or q in (None, ""):
            continue
        out.append(r)
    return out


def main(ruta, destino):
    wb = openpyxl.load_workbook(ruta, data_only=True)
    problemas = []

    # ── Catálogo ───────────────────────────────────────────────────────────
    inv = wb["INVENTARIO PARTES GASTABLES"]
    items = []
    por_codigo = {}
    sin_codigo = []

    # TODOS los códigos que aparecen en la hoja, incluidas las filas que se
    # descartan por no tener descripción. Si no se reservaran, el generador de
    # códigos nuevos podría reasignar uno que el Excel ya tenía apartado.
    usados = {
        txt(inv.cell(r, 1).value)
        for r in range(8, inv.max_row + 1)
        if txt(inv.cell(r, 1).value)
    }
    vistos = set()

    for r in range(8, inv.max_row + 1):
        vals = [inv.cell(r, c).value for c in range(1, 14)]
        if all(txt(v) in ("", "-") for v in vals):
            continue
        cod = txt(vals[0])
        desc_cruda = txt(vals[1])
        if not desc_cruda:
            problemas.append({"tipo": "ITEM_SIN_DESCRIPCION", "hoja": "INVENTARIO", "fila": r})
            continue
        desc, serie = extraer_serie(desc_cruda)
        item = {
            "fila": r,
            "codigo": cod or None,
            "descripcion": desc,
            "parte_no": txt(vals[2]) or None,
            "ubicacion": txt(vals[3]) or None,
            "categoria": clasificacion(vals[4]),
            "unidad": "UN",                       # se completa desde ENTRADAS
            "costo_unitario": vals[8] if isinstance(vals[8], (int, float)) else None,
            "serie_no": serie,
            "es_serializado": bool(serie),
            "stock_excel": vals[7] if isinstance(vals[7], (int, float)) else None,
            "clave_texto": (norm(desc_cruda), norm(vals[2])),
        }
        if cod and cod in vistos:
            problemas.append({
                "tipo": "CODIGO_DUPLICADO", "hoja": "INVENTARIO", "fila": r,
                "detalle": f"El código {cod} ya lo tiene otro ítem; a este se le asigna uno nuevo.",
            })
            item["codigo"] = None
        elif cod:
            vistos.add(cod)
        if not item["codigo"]:
            sin_codigo.append(item)
        items.append(item)

    # Códigos nuevos para los que no tenían (y para el duplicado), desde el
    # siguiente libre después del máximo del Excel.
    contador = [max((int(c) for c in usados if c.isdigit()), default=0) + 1]

    def nuevo_codigo():
        while str(contador[0]).zfill(6) in usados:
            contador[0] += 1
        c = str(contador[0]).zfill(6)
        usados.add(c)
        contador[0] += 1
        return c

    for it in sin_codigo:
        it["codigo"] = nuevo_codigo()
        problemas.append({
            "tipo": "CODIGO_ASIGNADO", "hoja": "INVENTARIO", "fila": it["fila"],
            "detalle": f"{it['descripcion']} no tenía código; se le asignó {it['codigo']}.",
        })

    for it in items:
        por_codigo[it["codigo"]] = it
    por_texto = defaultdict(list)
    for it in items:
        por_texto[it["clave_texto"]].append(it)

    # ── Movimientos ────────────────────────────────────────────────────────
    docs = {}

    def resolver(cod, desc, pn, hoja, fila, cant):
        """
        Código primero; si no resuelve, por texto (como hacía el Excel); y si
        tampoco, se DA DE ALTA el ítem con los datos del propio movimiento.

        Dar de alta es mejor que descartar: son materiales que de verdad
        entraron o salieron de la bodega y que nadie llegó a catalogar. Si se
        descartaran, el stock del resto quedaría bien pero esas piezas
        desaparecerían — exactamente el defecto del Excel que venimos a corregir.
        """
        c = txt(cod)
        if c and c in por_codigo:
            return por_codigo[c]
        clave = (norm(desc), norm(pn))
        cand = por_texto.get(clave, [])
        if len(cand) == 1:
            problemas.append({
                "tipo": "RESUELTO_POR_TEXTO", "hoja": hoja, "fila": fila,
                "detalle": f"'{txt(desc)}' (PN {txt(pn) or '—'}) no traía código válido; se cruzó por descripción con {cand[0]['codigo']}.",
            })
            return cand[0]

        d, serie = extraer_serie(desc)
        item = {
            "codigo": c if (c and c not in usados) else nuevo_codigo(),
            "descripcion": d, "parte_no": txt(pn) or None,
            "ubicacion": None, "categoria": None, "unidad": "UN",
            "costo_unitario": None, "serie_no": serie, "es_serializado": bool(serie),
            "stock_excel": None, "creado_desde_movimiento": True,
        }
        if c and c not in usados:
            usados.add(c)
        items.append(item)
        por_codigo[item["codigo"]] = item
        por_texto[clave].append(item)
        problemas.append({
            "tipo": "ITEM_CREADO_DESDE_MOVIMIENTO", "hoja": hoja, "fila": fila,
            "detalle": f"'{txt(desc)}' (PN {txt(pn) or '—'}, cant. {cant}) no estaba en la hoja de inventario; se dio de alta como {item['codigo']}. Revisar ubicación, clasificación y costo.",
        })
        return item

    def cargar(hoja, tipo, col_unidad=None):
        ws = wb[hoja]
        for r in filas(ws, 4, 6):
            doc_txt = txt(ws.cell(r, 1).value)
            fecha = ws.cell(r, 2).value
            cant, cant_texto = cantidad(ws.cell(r, 6).value)
            if cant is None:
                problemas.append({
                    "tipo": "CANTIDAD_ILEGIBLE", "hoja": hoja, "fila": r,
                    "detalle": f"Cantidad '{txt(ws.cell(r, 6).value)}' no tiene ningún número. NO se cargó.",
                })
                continue
            if cant_texto:
                problemas.append({
                    "tipo": "CANTIDAD_INTERPRETADA", "hoja": hoja, "fila": r,
                    "detalle": f"Cantidad '{cant_texto}' se interpretó como {cant}. Verificar (queda anotado en el renglón).",
                })

            m = re.match(r"^(FA|REQ)-(\d+)-(\d{4})$", doc_txt)
            if not m:
                problemas.append({
                    "tipo": "DOCUMENTO_ILEGIBLE", "hoja": hoja, "fila": r,
                    "detalle": f"Documento '{doc_txt}' no tiene el formato esperado. NO se cargó.",
                })
                continue
            numero = int(m.group(2))
            # El año sale de la FECHA, no del texto: así FA-00019-2027 (año mal
            # tecleado) cae en 2026 y se fusiona con su gemelo.
            anio = fecha.year if hasattr(fecha, "year") else int(m.group(3))
            if hasattr(fecha, "year") and int(m.group(3)) != fecha.year:
                problemas.append({
                    "tipo": "ANIO_CORREGIDO", "hoja": hoja, "fila": r,
                    "detalle": f"'{doc_txt}' dice {m.group(3)} pero la fecha es {fecha.date()}; se usa {fecha.year}.",
                })
            if len(m.group(2)) not in (3, 5):
                problemas.append({
                    "tipo": "CORRELATIVO_NORMALIZADO", "hoja": hoja, "fila": r,
                    "detalle": f"'{doc_txt}' tiene {len(m.group(2))} dígitos; se normaliza al número {numero}.",
                })

            item = resolver(ws.cell(r, 3).value, ws.cell(r, 4).value,
                            ws.cell(r, 5).value, hoja, r, cant)
            if not item:
                continue

            coment = txt(ws.cell(r, 7).value) if tipo == "SALIDA" else txt(ws.cell(r, 8).value)
            clave = (tipo, anio, numero)
            d = docs.get(clave)
            if not d:
                d = docs[clave] = {
                    "tipo": tipo, "anio": anio, "numero": numero,
                    "fecha": fecha.date().isoformat() if hasattr(fecha, "date") else None,
                    "motivo": coment or None,
                    "matricula": matricula(coment) if tipo == "SALIDA" else None,
                    "nota": coment if tipo == "ENTRADA" else None,
                    "renglones": [],
                }
            if d["fecha"] is None and hasattr(fecha, "date"):
                d["fecha"] = fecha.date().isoformat()
            nota_renglon = coment or None
            if cant_texto:
                nota_renglon = f"[cantidad original en el Excel: '{cant_texto}'] {nota_renglon or ''}".strip()
            d["renglones"].append({
                "codigo": item["codigo"], "cantidad": float(cant), "nota": nota_renglon,
            })

            if col_unidad:
                u = unidad(ws.cell(r, col_unidad).value)
                # La unidad del ítem sale de su ENTRADA más reciente.
                item["unidad"] = u

    cargar("ENTRADAS", "ENTRADA", col_unidad=7)
    cargar("SALIDAS", "SALIDA")

    for d in docs.values():
        if d["tipo"] == "SALIDA" and not d["matricula"]:
            problemas.append({
                "tipo": "SALIDA_SIN_AERONAVE", "hoja": "SALIDAS", "fila": None,
                "detalle": f"REQ-{d['numero']:03d}-{d['anio']}: el comentario no menciona matrícula. Queda sin aeronave (la obligatoriedad aplica solo hacia adelante).",
            })

    for it in items:
        it.pop("clave_texto", None)
        it.pop("fila", None)

    salida = {
        "items": items,
        "documentos": sorted(docs.values(), key=lambda d: (d["tipo"], d["anio"], d["numero"])),
        "externas": sorted(EXTERNAS),
        "problemas": problemas,
    }
    with open(destino, "w", encoding="utf-8") as f:
        json.dump(salida, f, ensure_ascii=False, indent=1)

    ent = [d for d in salida["documentos"] if d["tipo"] == "ENTRADA"]
    sal = [d for d in salida["documentos"] if d["tipo"] == "SALIDA"]
    print(f"items:      {len(items)}")
    print(f"entradas:   {len(ent)} documentos, {sum(len(d['renglones']) for d in ent)} renglones")
    print(f"salidas:    {len(sal)} documentos, {sum(len(d['renglones']) for d in sal)} renglones")
    print(f"problemas:  {len(problemas)}")
    for t in sorted({p["tipo"] for p in problemas}):
        print(f"   {t}: {sum(1 for p in problemas if p['tipo'] == t)}")
    print(f"-> {destino}")


if __name__ == "__main__":
    if len(sys.argv) < 3:
        sys.exit("uso: python extraer_excel.py <ruta.xlsx> <destino.json>")
    main(sys.argv[1], sys.argv[2])
